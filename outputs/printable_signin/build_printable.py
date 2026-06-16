from pathlib import Path
from math import ceil

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INPUT = Path("/Users/fushuo/Downloads/6.12监考员签到表.xlsx")
OUTPUT = Path("/Users/fushuo/Documents/分泌蛋白交互模型/outputs/printable_signin/6.12监考员签到表_打印版.xlsx")


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def collect_rows(ws, start_row, columns):
    rows = []
    for r in range(start_row, ws.max_row + 1):
        values = [cell_text(ws.cell(r, c).value) for c in columns]
        if not values[0]:
            continue
        rows.append(values)
    return rows


def split_two_pages(rows):
    pivot = ceil(len(rows) / 2)
    return rows[:pivot], rows[pivot:]


def setup_sheet(ws, title, headers, rows):
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15
    ws.print_options.horizontalCentered = True

    col_count = len(headers)
    last_col = get_column_letter(col_count)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws["A1"] = title
    ws["A1"].font = Font(name="宋体", size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.append(headers)
    for row in rows:
        ws.append(row)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=col_count):
        for cell in row:
            cell.border = border
            cell.font = Font(name="宋体", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[2]:
        cell.font = Font(name="宋体", size=11, bold=True)
        cell.fill = header_fill

    for idx in range(2, ws.max_row + 1):
        ws.row_dimensions[idx].height = 18.8

    widths = {
        5: [8, 13, 17, 16, 16],
        6: [8, 13, 14, 12, 16, 16],
    }[col_count]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.print_title_rows = "1:2"
    ws.print_area = f"A1:{last_col}{ws.max_row}"


def main():
    source_wb = load_workbook(INPUT, data_only=True)
    source_ws = source_wb.active

    jia_headers = ["序号", "监考甲", "甲单位", "甲联系电话", "签到"]
    yi_headers = ["序号", "监考乙", "考点", "乙单位", "乙联系电话", "签到"]
    jia_rows = collect_rows(source_ws, 5, [1, 2, 3, 4, 5])
    yi_rows = collect_rows(source_ws, 5, [7, 8, 9, 10, 11, 12])

    jia_1, jia_2 = split_two_pages(jia_rows)
    yi_1, yi_2 = split_two_pages(yi_rows)

    out_wb = Workbook()
    default = out_wb.active
    out_wb.remove(default)

    sheets = [
        ("甲第1页", "监考员甲签到表（第1页）", jia_headers, jia_1),
        ("甲第2页", "监考员甲签到表（第2页）", jia_headers, jia_2),
        ("乙第1页", "监考员乙签到表（第1页）", yi_headers, yi_1),
        ("乙第2页", "监考员乙签到表（第2页）", yi_headers, yi_2),
    ]
    for sheet_name, title, headers, rows in sheets:
        setup_sheet(out_wb.create_sheet(sheet_name), title, headers, rows)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(OUTPUT)
    print(OUTPUT)
    print(f"甲 {len(jia_rows)} 人，乙 {len(yi_rows)} 人；每类各拆成 {len(jia_1)} / {len(jia_2)} 行。")


if __name__ == "__main__":
    main()
